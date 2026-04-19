from fastapi import APIRouter, HTTPException, Depends, Response
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime, timezone
from database import db, logger
from auth import get_current_user
from uuid import uuid4

router = APIRouter()


class ArtikelLeistungCreate(BaseModel):
    name: str
    artikel_nr: str = ""
    description: str = ""
    typ: str = "Artikel"  # Artikel, Leistung, Fremdleistung
    price_net: float = 0
    ek_preis: float = 0
    aufschlag_1: float = 0
    aufschlag_2: float = 0
    aufschlag_3: float = 0
    vk_preis_1: float = 0
    vk_preis_2: float = 0
    vk_preis_3: float = 0
    unit: str = "Stk."
    subunternehmer: str = ""
    labor_cost: float = 0


class ArtikelLeistungUpdate(BaseModel):
    name: Optional[str] = None
    artikel_nr: Optional[str] = None
    description: Optional[str] = None
    typ: Optional[str] = None
    price_net: Optional[float] = None
    ek_preis: Optional[float] = None
    aufschlag_1: Optional[float] = None
    aufschlag_2: Optional[float] = None
    aufschlag_3: Optional[float] = None
    vk_preis_1: Optional[float] = None
    vk_preis_2: Optional[float] = None
    vk_preis_3: Optional[float] = None
    unit: Optional[str] = None
    subunternehmer: Optional[str] = None
    labor_cost: Optional[float] = None


ARTIKEL_MODUL = {
    "name": "Artikel & Leistungen",
    "slug": "artikel-leistungen",
    "version": "1.0.0",
    "description": "Eigenstaendiges Modul fuer Artikel, Leistungen und Fremdleistungen. Stellt Positionen fuer Angebote, Auftraege und Rechnungen bereit.",
    "status": "aktiv",
    "category": "daten",
    "data_collection": "module_artikel",
    "fields": [
        {"name": "artikel_nr", "type": "text", "label": "Artikel-Nr.", "required": False},
        {"name": "name", "type": "text", "label": "Bezeichnung", "required": True},
        {"name": "description", "type": "textarea", "label": "Beschreibung", "required": False},
        {"name": "typ", "type": "select", "label": "Typ", "options": ["Artikel", "Leistung", "Fremdleistung"], "required": True},
        {"name": "ek_preis", "type": "number", "label": "EK-Preis (Netto)", "required": False},
        {"name": "price_net", "type": "number", "label": "VK-Preis (Netto)", "required": True},
        {"name": "unit", "type": "select", "label": "Einheit", "options": ["Stk.", "Stunde", "m", "m2", "m3", "kg", "Psch.", "km"], "required": True},
        {"name": "subunternehmer", "type": "text", "label": "Subunternehmer", "required": False},
    ],
    "api_endpoints": [
        {"method": "GET", "path": "/api/modules/artikel/data", "description": "Alle Artikel & Leistungen abrufen"},
        {"method": "POST", "path": "/api/modules/artikel/data", "description": "Neuen Artikel/Leistung erstellen"},
        {"method": "PUT", "path": "/api/modules/artikel/data/{id}", "description": "Artikel/Leistung bearbeiten"},
        {"method": "DELETE", "path": "/api/modules/artikel/data/{id}", "description": "Artikel/Leistung loeschen"},
        {"method": "GET", "path": "/api/modules/artikel/export", "description": "Alle Daten exportieren"},
    ],
}


async def ensure_modul_registered():
    existing = await db.modules.find_one({"slug": "artikel-leistungen"})
    if not existing:
        from routes.modules import ModuleSchema
        modul = ModuleSchema(**ARTIKEL_MODUL)
        await db.modules.insert_one(modul.model_dump())
        logger.info("Artikel & Leistungen Modul registriert")


# ==================== KONFIGURATION ====================

DEFAULT_CONFIG = {
    "artikel_prefix": "ArtNr",
    "artikel_start": 2640,
    "leistung_prefix": "Leist",
    "leistung_start": 2660,
    "fremdleistung_prefix": "Fremd",
    "fremdleistung_start": 26000,
}


async def get_config():
    config = await db.module_artikel_config.find_one({"id": "config"}, {"_id": 0})
    if not config:
        config = {"id": "config", **DEFAULT_CONFIG}
        await db.module_artikel_config.insert_one(config)
        config.pop("_id", None)
    return config


async def generate_nummer(typ: str):
    """Generiert die naechste Nummer basierend auf Konfiguration"""
    config = await get_config()

    if typ == "Artikel":
        prefix = config.get("artikel_prefix", "ArtNr")
        start = config.get("artikel_start", 26000)
    elif typ == "Leistung":
        prefix = config.get("leistung_prefix", "Leist")
        start = config.get("leistung_start", 26000)
    else:  # Fremdleistung
        prefix = config.get("fremdleistung_prefix", "Fremd")
        start = config.get("fremdleistung_start", 26000)

    # Finde die hoechste bestehende Nummer fuer diesen Typ
    items = await db.module_artikel.find(
        {"typ": typ, "artikel_nr": {"$regex": f"^{prefix}"}},
        {"_id": 0, "artikel_nr": 1}
    ).to_list(10000)

    max_num = start - 1
    for item in items:
        try:
            num = int(item["artikel_nr"].replace(prefix, ""))
            if num > max_num:
                max_num = num
        except (ValueError, AttributeError):
            pass

    return f"{prefix}{max_num + 1}"


@router.get("/modules/artikel/config")
async def get_artikel_config(user=Depends(get_current_user)):
    return await get_config()


@router.put("/modules/artikel/config")
async def update_artikel_config(data: dict, user=Depends(get_current_user)):
    allowed = ["artikel_prefix", "artikel_start", "leistung_prefix", "leistung_start", "fremdleistung_prefix", "fremdleistung_start"]
    update = {k: v for k, v in data.items() if k in allowed}
    if not update:
        raise HTTPException(400, "Keine gueltige Konfiguration")
    await db.module_artikel_config.update_one({"id": "config"}, {"$set": update}, upsert=True)
    return await get_config()


@router.get("/modules/artikel/next-number/{typ}")
async def get_next_number(typ: str, user=Depends(get_current_user)):
    """Gibt die naechste verfuegbare Nummer zurueck"""
    if typ not in ["Artikel", "Leistung", "Fremdleistung"]:
        raise HTTPException(400, "Ungueltiger Typ")
    nummer = await generate_nummer(typ)
    return {"nummer": nummer}


@router.get("/modules/artikel/data")
async def get_artikel(user=Depends(get_current_user)):
    await ensure_modul_registered()
    items = await db.module_artikel.find({}, {"_id": 0}).sort("name", 1).to_list(10000)
    return items


@router.get("/modules/artikel/data/{item_id}")
async def get_artikel_item(item_id: str, user=Depends(get_current_user)):
    item = await db.module_artikel.find_one({"id": item_id}, {"_id": 0})
    if not item:
        raise HTTPException(404, "Nicht gefunden")
    return item


@router.post("/modules/artikel/data")
async def create_artikel(data: ArtikelLeistungCreate, user=Depends(get_current_user)):
    await ensure_modul_registered()
    # Auto-Nummer wenn leer
    artikel_nr = data.artikel_nr
    if not artikel_nr:
        artikel_nr = await generate_nummer(data.typ)
    item = {
        "id": str(uuid4()),
        **data.model_dump(),
        "artikel_nr": artikel_nr,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.module_artikel.insert_one(item)
    item.pop("_id", None)
    logger.info(f"Artikel erstellt: {artikel_nr} - {data.name} ({data.typ})")
    return item


@router.put("/modules/artikel/data/{item_id}")
async def update_artikel(item_id: str, data: ArtikelLeistungUpdate, user=Depends(get_current_user)):
    existing = await db.module_artikel.find_one({"id": item_id})
    if not existing:
        raise HTTPException(404, "Nicht gefunden")
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.module_artikel.update_one({"id": item_id}, {"$set": update_data})
    updated = await db.module_artikel.find_one({"id": item_id}, {"_id": 0})
    return updated


@router.delete("/modules/artikel/data/{item_id}")
async def delete_artikel(item_id: str, user=Depends(get_current_user)):
    result = await db.module_artikel.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Nicht gefunden")
    return {"message": "Geloescht"}


@router.get("/modules/artikel/export")
async def export_artikel(user=Depends(get_current_user)):
    items = await db.module_artikel.find({}, {"_id": 0}).to_list(10000)
    modul = await db.modules.find_one({"slug": "artikel-leistungen"}, {"_id": 0})
    return {"module": modul, "data": items, "exported_at": datetime.now(timezone.utc).isoformat(), "count": len(items)}


# ==================== DUPLIKAT-FINDER ====================

def _normalize_name(s: str) -> str:
    """Normalisiert Namen fuer Duplikat-Vergleich: lowercase, trim, collapse whitespace"""
    if not s:
        return ""
    import re
    s = s.lower().strip()
    s = re.sub(r"\s+", " ", s)
    # Entferne gaengige Fuellwoerter und Sonderzeichen
    s = re.sub(r"[.,;:!?()\[\]\-_/\\]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _completeness_score(item: dict) -> int:
    """Hoeher = vollstaendiger. Dient zur Empfehlung welcher Datensatz behalten werden soll."""
    score = 0
    if item.get("description"): score += 3
    if item.get("ek_preis", 0) > 0: score += 2
    if item.get("price_net", 0) > 0: score += 2
    if item.get("vk_preis_1", 0) > 0: score += 1
    if item.get("vk_preis_2", 0) > 0: score += 1
    if item.get("vk_preis_3", 0) > 0: score += 1
    if item.get("artikel_nr"): score += 2
    if item.get("unit"): score += 1
    if item.get("gewerk"): score += 1
    if item.get("lieferant"): score += 1
    if item.get("labor_cost", 0) > 0: score += 1
    return score


@router.get("/modules/artikel/find-duplicates")
async def find_duplicates(user=Depends(get_current_user)):
    """Findet Duplikate basierend auf normalisiertem Namen + gleichem Typ.
    Gibt Gruppen zurueck inkl. Empfehlung welcher behalten werden soll."""
    items = await db.module_artikel.find({}, {"_id": 0}).to_list(10000)

    # Gruppiere nach (typ, normalisierter_name)
    groups = {}
    for it in items:
        key = (it.get("typ", "Artikel"), _normalize_name(it.get("name", "")))
        if not key[1]:
            continue
        groups.setdefault(key, []).append(it)

    duplicates = []
    for (typ, norm_name), group in groups.items():
        if len(group) < 2:
            continue
        # Sort: hoechster Completeness-Score zuerst, bei Gleichstand aeltester (created_at) zuerst
        scored = [(_completeness_score(i), -1 * len(i.get("created_at", "")), i) for i in group]
        scored.sort(key=lambda x: (-x[0], x[1]))

        keep = scored[0][2]
        delete_candidates = [s[2] for s in scored[1:]]

        duplicates.append({
            "typ": typ,
            "normalized_name": norm_name,
            "count": len(group),
            "keep_recommended": {
                "id": keep.get("id"),
                "name": keep.get("name"),
                "artikel_nr": keep.get("artikel_nr", ""),
                "price_net": keep.get("price_net", 0),
                "ek_preis": keep.get("ek_preis", 0),
                "unit": keep.get("unit", ""),
                "description": keep.get("description", ""),
                "score": _completeness_score(keep),
            },
            "delete_suggestions": [
                {
                    "id": i.get("id"),
                    "name": i.get("name"),
                    "artikel_nr": i.get("artikel_nr", ""),
                    "price_net": i.get("price_net", 0),
                    "ek_preis": i.get("ek_preis", 0),
                    "unit": i.get("unit", ""),
                    "description": i.get("description", ""),
                    "score": _completeness_score(i),
                }
                for i in delete_candidates
            ],
        })

    # Sortiere Gruppen: groesste zuerst
    duplicates.sort(key=lambda g: -g["count"])
    return {
        "total_items": len(items),
        "duplicate_groups": len(duplicates),
        "total_duplicates_found": sum(g["count"] - 1 for g in duplicates),
        "groups": duplicates,
    }


@router.post("/modules/artikel/bulk-delete")
async def bulk_delete(body: dict, user=Depends(get_current_user)):
    """Loescht eine Liste von Artikel-IDs auf einmal"""
    ids = body.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "Keine IDs angegeben")
    result = await db.module_artikel.delete_many({"id": {"$in": ids}})
    logger.info(f"Bulk-delete: {result.deleted_count} Artikel geloescht")
    return {"deleted": result.deleted_count, "requested": len(ids)}



# ==================== CSV IMPORT ====================

@router.get("/modules/artikel/import-vorlage")
async def download_vorlage():
    """CSV-Vorlage herunterladen"""
    import csv, io
    output = io.StringIO()
    w = csv.writer(output, delimiter=";")
    w.writerow(["Typ", "Artikel_Nr", "Name", "Beschreibung", "Einheit", "EK_Preis", "VK_Preis_1", "VK_Preis_2", "VK_Preis_3", "Gewerk", "Lieferant", "Subunternehmer"])
    w.writerow(["Artikel", "", "Beispiel Material", "Beschreibung hier", "Stk.", "10.00", "15.00", "14.00", "13.00", "Sanitaer", "", "Nein"])
    w.writerow(["Leistung", "", "Beispiel Arbeitszeit", "Stundensatz Monteur", "Stunde", "0", "65.00", "60.00", "55.00", "", "", "Nein"])
    return Response(content=output.getvalue().encode("utf-8-sig"), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=Vorlage_Artikel_Import.csv"})


@router.get("/modules/artikel/download-alle")
async def download_alle_artikel():
    """Alle Artikel als CSV herunterladen (ohne Auth)"""
    import os
    from fastapi.responses import FileResponse
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "alle_artikel_zum_import.csv")
    if not os.path.exists(path):
        raise HTTPException(404, "Datei nicht gefunden")
    return FileResponse(path, filename="Alle_Artikel_Graupner.csv", media_type="text/csv",
                        headers={"Content-Disposition": "attachment; filename=Alle_Artikel_Graupner.csv"})


@router.get("/modules/artikel/export-material-csv")
async def export_material_csv():
    """Export Material CSV aus alter Datenbank"""
    import os
    from fastapi.responses import FileResponse
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "export_material.csv")
    if not os.path.exists(path):
        raise HTTPException(404, "Datei nicht gefunden")
    return FileResponse(path, filename="Material_aus_Altdatenbank.csv", media_type="text/csv")


@router.get("/modules/artikel/export-leistungen-csv")
async def export_leistungen_csv():
    """Export Leistungen CSV aus alter Datenbank"""
    import os
    from fastapi.responses import FileResponse
    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "export_leistungen.csv")
    if not os.path.exists(path):
        raise HTTPException(404, "Datei nicht gefunden")
    return FileResponse(path, filename="Leistungen_aus_Altdatenbank.csv", media_type="text/csv")


# ==================== IMPORT / EXPORT MODUL ====================

from fastapi import UploadFile, File as FileParam

@router.post("/modules/artikel/import-upload")
async def import_upload(file: UploadFile = FileParam(...), user=Depends(get_current_user)):
    """CSV, Excel, JSON oder XML importieren mit Duplikat-Pruefung und Auto-Nummern"""
    content = await file.read()
    filename = (file.filename or "").lower()

    rows = []
    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            rows = _parse_excel(content)
        elif filename.endswith(".json"):
            rows = _parse_json(content)
        elif filename.endswith(".xml"):
            rows = _parse_xml(content)
        else:
            rows = _parse_csv(content)
    except Exception as e:
        raise HTTPException(400, f"Datei konnte nicht gelesen werden: {str(e)}")

    if not rows:
        raise HTTPException(400, "Keine Daten gefunden")

    imported = 0
    skipped = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()

    for i, row in enumerate(rows):
        try:
            name = (row.get("Name") or row.get("name") or row.get("Kurztext") or "").strip()
            if not name:
                skipped += 1
                continue

            typ = (row.get("Typ") or row.get("typ") or "Artikel").strip()
            if typ.lower() in ("leistung", "service", "dienstleistung"):
                typ = "Leistung"
            elif typ.lower() in ("personal", "lohn"):
                typ = "Leistung"
            elif typ.lower() in ("fremdleistung", "fremd"):
                typ = "Fremdleistung"
            else:
                typ = "Artikel"

            # Duplikat-Pruefung: Name bereits vorhanden?
            existing = await db.module_artikel.find_one({"name": name})
            if existing:
                skipped += 1
                continue

            # Auto-Nummer wenn nicht angegeben
            artikel_nr = (row.get("Artikel_Nr") or row.get("artikel_nr") or row.get("Nr") or "").strip()
            if not artikel_nr:
                artikel_nr = await generate_nummer(typ)

            item = {
                "id": str(uuid4()),
                "name": name,
                "artikel_nr": artikel_nr,
                "description": (row.get("Beschreibung") or row.get("description") or row.get("Langtext") or "").strip(),
                "typ": typ,
                "unit": (row.get("Einheit") or row.get("unit") or "Stk.").strip(),
                "ek_preis": _to_float(row.get("EK_Preis") or row.get("ek_preis")),
                "price_net": _to_float(row.get("VK_Preis_1") or row.get("vk_preis_1") or row.get("price_net") or row.get("EK_Preis") or row.get("ek_preis")),
                "vk_preis_1": _to_float(row.get("VK_Preis_1") or row.get("vk_preis_1")),
                "vk_preis_2": _to_float(row.get("VK_Preis_2") or row.get("vk_preis_2")),
                "vk_preis_3": _to_float(row.get("VK_Preis_3") or row.get("vk_preis_3")),
                "aufschlag_1": _to_float(row.get("Aufschlag_1") or row.get("aufschlag_1")),
                "aufschlag_2": _to_float(row.get("Aufschlag_2") or row.get("aufschlag_2")),
                "aufschlag_3": _to_float(row.get("Aufschlag_3") or row.get("aufschlag_3")),
                "gewerk": (row.get("Gewerk") or row.get("gewerk") or "").strip(),
                "lieferant": (row.get("Lieferant") or row.get("lieferant") or "").strip(),
                "subunternehmer": (row.get("Subunternehmer") or "").strip().lower() in ("ja", "yes", "true", "1"),
                "created_at": now,
                "updated_at": now,
            }
            await db.module_artikel.insert_one(item)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i+2}: {str(e)}")

    return {
        "message": f"{imported} importiert, {skipped} uebersprungen (Duplikate/leer)",
        "imported": imported,
        "skipped": skipped,
        "errors": errors
    }


# ---- Export Formate ----

@router.get("/modules/artikel/export-csv")
async def export_csv(user=Depends(get_current_user)):
    """Alle Artikel als CSV exportieren"""
    import csv, io
    items = await db.module_artikel.find({}, {"_id": 0}).sort("typ", 1).to_list(10000)
    output = io.StringIO()
    w = csv.writer(output, delimiter=";")
    w.writerow(["Typ", "Artikel_Nr", "Name", "Beschreibung", "Einheit", "EK_Preis", "VK_Preis_1", "VK_Preis_2", "VK_Preis_3", "Gewerk", "Lieferant", "Subunternehmer"])
    for a in items:
        w.writerow([a.get("typ",""), a.get("artikel_nr",""), a.get("name",""), (a.get("description","") or "").replace("\n"," "), a.get("unit",""), a.get("ek_preis",0), a.get("vk_preis_1") or a.get("price_net",0), a.get("vk_preis_2",0), a.get("vk_preis_3",0), a.get("gewerk",""), a.get("lieferant",""), "Ja" if a.get("subunternehmer") else "Nein"])
    return Response(content=output.getvalue().encode("utf-8-sig"), media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=Artikel_Export.csv"})


@router.get("/modules/artikel/export-excel")
async def export_excel(user=Depends(get_current_user)):
    """Alle Artikel als Excel exportieren"""
    import openpyxl, io
    items = await db.module_artikel.find({}, {"_id": 0}).sort("typ", 1).to_list(10000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artikel & Leistungen"
    headers = ["Typ", "Artikel_Nr", "Name", "Beschreibung", "Einheit", "EK_Preis", "VK_Preis_1", "VK_Preis_2", "VK_Preis_3", "Gewerk", "Lieferant", "Subunternehmer"]
    ws.append(headers)
    # Header fett
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for a in items:
        ws.append([a.get("typ",""), a.get("artikel_nr",""), a.get("name",""), (a.get("description","") or "").replace("\n"," "), a.get("unit",""), a.get("ek_preis",0), a.get("vk_preis_1") or a.get("price_net",0), a.get("vk_preis_2",0), a.get("vk_preis_3",0), a.get("gewerk",""), a.get("lieferant",""), "Ja" if a.get("subunternehmer") else "Nein"])
    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=Artikel_Export.xlsx"})


@router.get("/modules/artikel/export-json")
async def export_json(user=Depends(get_current_user)):
    """Alle Artikel als JSON exportieren"""
    import json as _json
    items = await db.module_artikel.find({}, {"_id": 0}).sort("typ", 1).to_list(10000)
    return Response(content=_json.dumps(items, indent=2, ensure_ascii=False).encode("utf-8"),
                    media_type="application/json",
                    headers={"Content-Disposition": "attachment; filename=Artikel_Export.json"})


@router.get("/modules/artikel/export-xml")
async def export_xml(user=Depends(get_current_user)):
    """Alle Artikel als XML exportieren"""
    from xml.sax.saxutils import escape
    items = await db.module_artikel.find({}, {"_id": 0}).sort("typ", 1).to_list(10000)
    lines = ['<?xml version="1.0" encoding="UTF-8"?>', "<artikel_export>"]
    for a in items:
        lines.append("  <item>")
        for key in ["typ", "artikel_nr", "name", "description", "unit", "ek_preis",
                    "vk_preis_1", "vk_preis_2", "vk_preis_3", "price_net",
                    "gewerk", "lieferant", "subunternehmer"]:
            val = a.get(key, "")
            if val is None:
                val = ""
            lines.append(f"    <{key}>{escape(str(val))}</{key}>")
        lines.append("  </item>")
    lines.append("</artikel_export>")
    xml_content = "\n".join(lines)
    return Response(content=xml_content.encode("utf-8"),
                    media_type="application/xml",
                    headers={"Content-Disposition": "attachment; filename=Artikel_Export.xml"})


# ---- Hilfsfunktionen ----

def _to_float(val):
    if not val:
        return 0.0
    try:
        return float(str(val).replace(",", ".").replace(" ", "").strip())
    except:
        return 0.0


def _parse_csv(content: bytes) -> list:
    import csv, io
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            text = content.decode(enc)
            break
        except:
            continue
    else:
        return []
    delimiter = ";" if ";" in text.split("\n")[0] else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return list(reader)


def _parse_excel(content: bytes) -> list:
    import openpyxl, io
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = str(val) if val is not None else ""
        result.append(d)
    return result


def _parse_json(content: bytes) -> list:
    import json as _json
    data = _json.loads(content.decode("utf-8"))
    # Accept top-level list, or {"data": [...]}, or {"items": [...]}
    if isinstance(data, dict):
        for k in ("data", "items", "artikel", "records"):
            if k in data and isinstance(data[k], list):
                return data[k]
        return [data]
    if isinstance(data, list):
        return data
    return []


def _parse_xml(content: bytes) -> list:
    import xml.etree.ElementTree as ET
    root = ET.fromstring(content.decode("utf-8"))
    result = []
    # Find item-like children (item/record/artikel)
    for child in root:
        d = {}
        for sub in child:
            d[sub.tag] = (sub.text or "").strip()
        if d:
            result.append(d)
    return result

